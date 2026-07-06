# excel_generator.py — Genere un .xlsx (tableau + totaux) à partir du dict facture.
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E4057")

def init_workbook():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Factures"
    wb.create_sheet("Clients")
    wb.create_sheet("Produits")
    return wb

def write_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

def json_vers_excel(facture: dict):
    f = facture.get("facture", {})
    c = facture.get("client", {})
    produits = facture.get("produits", [])

    # Génération d'un nom de fichier unique basé sur le numéro et la date de facture
    def nettoyer(s):
        """Remplace les caractères interdits dans les noms de fichiers."""
        return "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in str(s)).strip("-")

    num   = nettoyer(f.get("numero_facture", "")) or "inconnu"
    date  = nettoyer(f.get("date", ""))           or "sans-date"

    filename = f"facture_{num}_{date}.xlsx"
    filepath = f"{OUTPUT_DIR}/{filename}"

    wb = init_workbook()
    ws_fac = wb["Factures"]
    ws_cli = wb["Clients"]
    ws_pro = wb["Produits"]

    # Injecter la devise dans l'objet facture pour l'écrire dans l'Excel
    analyse = facture.get("analyse", {})
    f["devise"] = (analyse.get("devise_detecte") or "TND").strip()

    # --- Feuille Factures ---
    fac_headers = [
        "numero_facture", "date", "societe_nom", "societe_adresse",
        "societe_tel", "societe_email", "total_ht", "montant_tva",
        "timbre_fiscal", "net_a_payer", "mode_reglement", "etat", "remarques", "devise"
    ]
    
    # Intégrer les champs supplémentaires dynamiques
    champs_sup = f.get("champs_supplementaires") or {}
    champs_sup_keys = [k for k, v in champs_sup.items() if v not in (None, "", -9999, [], {})]
    
    # Filtrer les colonnes vides (null ou -9999) pour cette facture
    fac_headers = [h for h in fac_headers if f.get(h) not in (None, "", -9999)]
    
    # Ajouter les clés supplémentaires à la fin
    fac_headers.extend(champs_sup_keys)
    
    # Colonne "remarque" obligatoire en dernier
    fac_headers.append("remarque")
    
    write_headers(ws_fac, fac_headers)
    # Ecrire les valeurs en ligne 2
    for col, h in enumerate(fac_headers, 1):
        if h == "remarque":
            parties = []
            if f.get("mode_reglement"):
                parties.append(f"Paiement : {f['mode_reglement']}")
            if f.get("etat"):
                parties.append(f"État : {f['etat']}")
            echeance = champs_sup.get("Echeance") or champs_sup.get("Échéance") or champs_sup.get("echeance")
            if echeance:
                parties.append(f"Échéance : {echeance}")
            if f.get("remarques"):
                parties.append(f"{f['remarques']}")
            ws_fac.cell(2, col, "; ".join(parties) if parties else "Rien à signaler")
        elif h in champs_sup_keys:
            val = champs_sup.get(h)
            if isinstance(val, (dict, list)):
                val = str(val)
            ws_fac.cell(2, col, val)
        else:
            ws_fac.cell(2, col, f.get(h))

    # --- Feuille Clients ---
    cli_headers = [
        "numero_facture", "code_client", "nom", "prenom",
        "telephone", "adresse", "matricule_fiscal"
    ]
    c["numero_facture"] = f.get("numero_facture")
    # Filtrer les colonnes vides pour ce client
    cli_headers = [h for h in cli_headers if c.get(h) not in (None, "", -9999)]
    # Ajouter les champs supplementaires client
    cli_champs_sup = c.get("champs_supplementaires") or {}
    cli_champs_sup_keys = [k for k, v in cli_champs_sup.items() if v not in (None, "", -9999, [], {})]
    cli_headers.extend(cli_champs_sup_keys)
    write_headers(ws_cli, cli_headers)
    for col, h in enumerate(cli_headers, 1):
        if h in cli_champs_sup_keys:
            val = cli_champs_sup.get(h)
            if isinstance(val, (dict, list)):
                val = str(val)
            ws_cli.cell(2, col, val)
        else:
            ws_cli.cell(2, col, c.get(h))

    # --- Feuille Produits ---
    col_order = facture.get("col_order", [])
    if col_order:
        pro_headers = [c["label"] for c in col_order]
    else:
        std_cols = [
            ("designation", "Désignation"), ("quantite", "Qté"),
            ("prix_u_ht", "Prix U HT"), ("tva_pct", "TVA %"),
            ("remise_pct", "Remise %"), ("total_ht_ligne", "Total HT"),
            ("total_ttc", "Total TTC"),
        ]
        pro_headers = []
        for key, label in std_cols:
            if any(p.get(key) not in (None, "", -9999) for p in produits):
                pro_headers.append(label)
        pro_champs_sup_keys = []
        for p in produits:
            sup = p.get("champs_supplementaires") or {}
            for k in sup:
                if k not in pro_champs_sup_keys:
                    pro_champs_sup_keys.append(k)
        pro_headers.extend(pro_champs_sup_keys)

    write_headers(ws_pro, pro_headers)
    for row_idx, p in enumerate(produits, 2):
        pro_champs_sup = p.get("champs_supplementaires") or {}
        for col, col_def in enumerate(col_order, 1):
            if col_def.get("field"):
                val = p.get(col_def["field"])
            else:
                val = pro_champs_sup.get(col_def["supp_key"])
            if isinstance(val, (dict, list)):
                val = str(val)
            ws_pro.cell(row_idx, col, val if val not in (-9999,) else None)

    wb.save(filepath)
    return filename